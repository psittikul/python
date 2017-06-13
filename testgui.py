import sys
import os
import win32com.client
import errno
import tkinter
import openpyxl
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from flask import Flask, render_template, request
from werkzeug import secure_filename
import tkinter as tk
from tkinter import Tk
from tkinter import filedialog
from tkinter.filedialog import askopenfilename
from collections import defaultdict
UPLOAD_FOLDER = os.getcwd()
ALLOWED_EXTENSIONS = set({'xls', 'xlsx'})
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

     
class Example(QWidget):
    
    # --- Function to prompt user to upload the file with the updated objectives, and then collect and store the relevant data in an array, which will later be prepped to mail merge. ---
    # !! Add code in to make sure it's an accepted file type later
##    def allowed_file(filename):
##        return '.' in filename and \
##               filename.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS


    # Function to actually get the objectives update data
    def objectives(self):

        # Internal helper function to format the currency values with $ and comma separation
        def formatCurrency(amt):
            if len(amt) > 3:
                hundreds = amt[len(amt)-3:len(amt)+1]
                thousands = amt[:len(amt)-3]
                return "$"+thousands+","+hundreds
            else:
                return "$"+amt
        # -----------------------------------------------------------------------------------

        print("Getting file")
        Tk().withdraw()
        self.filename = askopenfilename()
        if not self.filename:
            print("Cancel pressed")
            self.close()

        print("Uploading to " + str(UPLOAD_FOLDER))
        data_wb = openpyxl.load_workbook(self.filename)
        print("Opening this uploaded file")
        # Go to the correct worksheet
        sheet = data_wb.active
        print(sheet)
        # Before we go through any file, create a Mail Merge workbook if it doesn't already exist, and just open it to update if it does
        print("Checking if Mail Merge workbook exists or not")
        mergeFilePath = str(os.getcwd() + "\MailMerge.xlsx")
        print(mergeFilePath)
        if os.path.exists(mergeFilePath):
            print("Opening existing Mail Merge workbook")
            merge_wb = openpyxl.load_workbook('MailMerge.xlsx')
            merge_sheet = merge_wb.active
        else:
            # if the file doesn't exist yet, create a new mail merge file
            merge_wb = openpyxl.Workbook()
            merge_sheet = merge_wb.active
            print("Adding column headers")
            merge_sheet['A1'] = "OSC Number"
            merge_sheet['B1'] = "DBA"
            merge_sheet['C1'] = "Contact First Name"
            merge_sheet['D1'] = "Contact Last Name"
            merge_sheet['E1'] = "Contact Email"
            merge_sheet['F1'] = "Aero Status"
            merge_sheet['G1'] = "Objective"
            merge_sheet['H1'] = "MTD"
            merge_sheet['I1'] = "% of Goal"
            merge_sheet['J1'] = "Purchases ToGo"
            merge_sheet['K1'] = "Reward"
            print("All headers added")
            merge_wb.save('MailMerge.xlsx')

        # Reactivate the objectives update workbook
        print("Reactivating objectives workbook now")
        sheet = data_wb.active
        # Cycle through each row. If the row is an OSC, gather this data to later be stored in the array "objectives_update"
        print("Entering for loop to gather data")
        # We'll start populating the mail merge worksheet at row 2
        mergeRow = 2
        # Start at top of file looking for the column headers (taking note of index) to determine the location of each field
        # Then gather those values to be stored in variables for copying over to "Mail Merge"
        for row in range(23, 30):
            for column in range(1, sheet.max_column + 1):
                if sheet.cell(row=row, column=column).value == "Rwd.":
                    rewardColumn = column
                    rewardLabel = "Rwd."
                    labelRow = row
                    break
                else:
                    continue
        codeLabel = str(sheet.cell(row=labelRow, column=1).value)
        aeroColumn = rewardColumn - 1
        aeroLabel = str(sheet.cell(row=labelRow, column=aeroColumn).value)
        DBAColumn = rewardColumn + 2
        DBALabel = str(sheet.cell(row=labelRow, column=DBAColumn).value)
        objectiveColumn = DBAColumn + 2
        objectiveLabel = str(sheet.cell(row=labelRow, column=objectiveColumn).value)
        PTDColumn = objectiveColumn + 1
        PTDLabel = str(sheet.cell(row=labelRow, column=PTDColumn).value)
        percentColumn = PTDColumn + 1
        percentLabel = str(sheet.cell(row=labelRow, column=percentColumn).value)
        PTGColumn = percentColumn + 3
        PTGLabel = str(sheet.cell(row=labelRow, column=PTGColumn).value)

        # Now that you've determined the location of each field, go through every row in the sheet to get actual data
        # --------------------------------------------------------------------------------------------------------------------------
        for row in range(23, sheet.max_row + 1):
            if str(sheet.cell(row=row, column=2).value) == "OSC" and type(
                    sheet.cell(row=row, column=rewardColumn).value) == int:
                code = str(sheet.cell(row=row, column=1).value)
                # If the OSC code is negative, take the "-" symbol out of it
                if "-" in code:
                    code = code[1:len(code) + 1]
                else:
                    code = code
                # Reward will be in the column labeled "Rwd."
                rewardCell = sheet.cell(row=row, column=rewardColumn)
                reward = "$" + str(rewardCell.value)
                # Aero status will be one cell to the left of reward
                aeroCell = sheet.cell(row=row, column=aeroColumn)
                aeroStatus = str(aeroCell.value)
                # DBA will be 2 cells to the right of reward
                DBACell = sheet.cell(row=row, column=DBAColumn)
                DBA = str(DBACell.value)
                # Objective will be 2 cells to the right of DBA
                print("Getting objective value")
                objectiveCell = sheet.cell(row=row, column=objectiveColumn)
                print(objectiveCell.value)
                objective = str(objectiveCell.value)
                objective = formatCurrency(objective)
                # Purchases to date will be 1 cell to the right of objective
                PTDCell = sheet.cell(row=row, column=PTDColumn)
                PTD = str(PTDCell.value)
                PTD = formatCurrency(PTD)
                # % of objective will be 1 cell to the right of PTD
                percentCell = sheet.cell(row=row, column=percentColumn)
                print(percentCell.value)
                percent = int(percentCell.value * 100)
                print(percent)
                percent = str(percent) + "%"
                # Purchases to go will be 3 cells to the right of % of objective
                PTGCell = sheet.cell(row=row, column=PTGColumn)
                PTG = str(PTGCell.value)
                PTG = formatCurrency(PTG)
                print(code + " " + DBA + " " + aeroStatus + " " + str(objective) + " " + reward)
                # Now that you've gotten the information, copy it over to Mail Merge
                merge_sheet = merge_wb.active
                # Go through all the info columns in Mail Merge and copy over the values
                merge_sheet.cell(row=mergeRow, column=1).value = code
                merge_sheet.cell(row=mergeRow, column=2).value = DBA
                merge_sheet.cell(row=mergeRow, column=6).value = aeroStatus
                merge_sheet.cell(row=mergeRow, column=7).value = objective
                merge_sheet.cell(row=mergeRow, column=8).value = PTD
                merge_sheet.cell(row=mergeRow, column=9).value = percent
                merge_sheet.cell(row=mergeRow, column=10).value = PTG
                merge_sheet.cell(row=mergeRow, column=11).value = reward
                # Now that we've copied the information to this row, get the next merge row ready and reactivate the objectives workbook and worksheet
                merge_wb.save("MailMerge.xlsx")
                mergeRow = mergeRow + 1
                sheet = data_wb.active
            else:
                continue
            merge_sheet = merge_wb.active
            merge_wb.save("MailMerge.xlsx")
        print(merge_sheet.cell(row=2, column=5).value)
        if merge_sheet.cell(row=2, column=5).value is None:
            # Ask user if they'd like to finish and view file, finish and exit, or upload contacts
            # Only call self.empty() on either finish buttons
            print("Getting emails")
            self.emails()
        else:
            self.close()

    # Function to ask for OSC Master upload and copy those emails to the Mail Merge file
    def emails(self):
        # Helper function to populate the first column of MailMerge.xlsx with the OSC codes
        def codes(code_column):
            OSC_sheet = OSC_wb.active
            mergeRow = 2
            for row in range(2, OSC_sheet.max_row + 1):
                code = str(OSC_sheet.cell(row = row, column = code_column).value)
                merge_sheet = merge_wb.active
                merge_sheet.cell(row = mergeRow, column = 1).value = code
                mergeRow += 1
                merge_wb.save("MailMerge.xlsx")
                OSC_sheet = OSC_wb.active
        #------------------------------------------------------------------------------------

        print("Getting file")
        Tk().withdraw()
        self.filename = askopenfilename()
        if not self.filename:
            print("Cancel pressed")
            self.close()
        print("Uploading to " + str(UPLOAD_FOLDER))
        OSC_wb = openpyxl.load_workbook(self.filename)
        print("Opening this uploaded file")
        # Go to the correct worksheet
        OSC_sheet = OSC_wb.active
        # Before we go through any file, create a Mail Merge workbook if it doesn't already exist, and just open it to update if it does
        print("Checking if Mail Merge workbook exists or not")
        mergeFilePath = str(os.getcwd() + "\MailMerge.xlsx")
        print(mergeFilePath)
        if os.path.exists(mergeFilePath):
            print("Opening existing Mail Merge workbook")
            merge_wb = openpyxl.load_workbook('MailMerge.xlsx')
            merge_sheet = merge_wb.active
            # Clear old values of the MailMerge sheet
            for row in range(2, merge_sheet.max_row + 1):
                for col in range(1, merge_sheet.max_column + 1):
                    merge_sheet.cell(row = row, column = col).value = None
            merge_wb.save("MailMerge.xlsx")
        else:
            # if the file doesn't exist yet, create a new mail merge file
            merge_wb = openpyxl.Workbook()
            merge_sheet = merge_wb.active
            print("Adding column headers")
            merge_sheet['A1'] = "OSC Number"
            merge_sheet['B1'] = "DBA"
            merge_sheet['C1'] = "Contact First Name"
            merge_sheet['D1'] = "Contact Last Name"
            merge_sheet['E1'] = "Contact Email"
            merge_sheet['F1'] = "Aero Status"
            merge_sheet['G1'] = "Objective"
            merge_sheet['H1'] = "MTD"
            merge_sheet['I1'] = "% of Goal"
            merge_sheet['J1'] = "Purchases ToGo"
            merge_sheet['K1'] = "Reward"
            print("All headers added")
            merge_wb.save("MailMerge.xlsx")

        # Reactivate the OSC Master workbook
        print("Reactivating OSC Master workbook now")
        OSC_sheet = OSC_wb.active
        # Cycle through each row. If the row is an OSC, gather this data to later be stored in the array "objectives_update"
        print("Entering for loop to gather data")
        # We'll start populating the mail merge worksheet at row 2
        mergeRow = 2
        print(OSC_sheet)
        # Start at header row of OSC Master to determine the column each manager's email is in
        for column in range(1, OSC_sheet.max_column + 1):
            print(str(OSC_sheet.cell(row = 1, column = column).value))
            if str(OSC_sheet.cell(row = 1, column = column).value) == "OSC":
                code_column = column
                print("OSC code column = " + str(code_column))
            if str(OSC_sheet.cell(row=1, column=column).value) == "General Manager Name":
                print("GM Name")
                GM_name_column = column
                print("General Manager Name column = " + str(GM_name_column))
            if str(OSC_sheet.cell(row = 1, column = column).value) == "Email (1)":
                GM_email_column = column
                print("General Manager Email column = " + str(GM_email_column))
            if str(OSC_sheet.cell(row = 1, column = column).value) == "Parts Manager Name":
                PM_name_column = column
                print("Parts Manager Name column = " + str(PM_name_column))
            if str(OSC_sheet.cell(row = 1, column = column).value) =="Email (2)":
                PM_email_column = column
                print("Parts Manager Email column = " + str(PM_email_column))
            if str(OSC_sheet.cell(row = 1, column = column).value) == "Service Manager Name":
                SM_name_column = column
                print("Service Manager Name column = " + str(SM_name_column))
            if str(OSC_sheet.cell(row = 1, column = column).value) == "Email (3)":
                SM_email_column = column
                print("Service Manager Email column = " + str(SM_email_column))
                break
        # Now that you've determined the location of each field, go through every row in the sheet to get contact info
        # But skip any emails that have already been copied or are blank
        codes(code_column)
        for row in range(2, OSC_sheet.max_row + 1):
            print("Getting emails")
            if "Active" in str(OSC_sheet.cell(row = row, column = 1).value):
                OSC_names = []
                OSC_emails = []
                PM_email = str(OSC_sheet.cell(row=row, column=PM_email_column).value)
                print("PM Email: " + PM_email)
                GM_email = str(OSC_sheet.cell(row=row, column=GM_email_column).value)
                print("GM Email: " + GM_email)
                SM_email = str(OSC_sheet.cell(row=row, column=SM_email_column).value)
                print("SM Email: " + SM_email)
                code = str(OSC_sheet.cell(row = row, column = code_column).value)
                print(code)
                if "@" in GM_email:
                    GM_name = str(OSC_sheet.cell(row = row, column = GM_name_column).value)
                    print("GM Name: " + GM_name)
                    OSC_names.append(GM_name)
                    OSC_emails.append(GM_email)
                if "@" in PM_email and PM_email != GM_email and PM_email != SM_email:
                    PM_name = str(OSC_sheet.cell(row = row, column = PM_name_column).value)
                    print("PM Name: " + PM_name)
                    OSC_names.append(PM_name)
                    OSC_emails.append(PM_email)
                if "@" in SM_email and SM_email != PM_email and SM_email != GM_email:
                    SM_name = str(OSC_sheet.cell(row = row, column = SM_name_column).value)
                    print("SM Name: " + SM_name)
                    OSC_names.append(SM_name)
                    OSC_emails.append(SM_email)
            else:
                continue
            # Now that you've gotten the information, copy it over to Mail Merge
            merge_sheet = merge_wb.active
            for row in range(2, merge_sheet.max_row + 1):
                if str(merge_sheet.cell(row = row, column = 1).value) == code:
                    print("OSC #: " + code + "; GM: " + OSC_names[0])
                    name = OSC_names[0]
                    print(str(OSC_names[0]))
                    fn = name[:name.rindex(" ")]
                    ln = name[name.rindex(" ") + 1:]
                    email = OSC_emails[0]
                    print(fn)
                    merge_sheet.cell(row = row, column = 3).value = fn
                    merge_sheet.cell(row = row, column = 4).value = ln
                    merge_sheet.cell(row = row, column = 5).value = email
                    if len(OSC_emails) > 1:
                        print(len(OSC_emails))
                        print("PM Name: " + OSC_names[1])
                        for em in OSC_emails:
                            print(OSC_emails.index(email))
                            curr_index = OSC_emails.index(em) + 1
                            if curr_index > len(OSC_emails) - 1:
                                print("Reached end of list")
                                break
                            newRow = merge_sheet.max_row + 1
                            name = OSC_names[curr_index]
                            print(name)
                            fn = name[:name.rindex(" ")]
                            ln = name[name.rindex(" ") + 1:]
                            email = OSC_emails[curr_index]
                            merge_sheet.cell(row = newRow, column = 1).value = code
                            merge_sheet.cell(row = newRow, column = 2).value = merge_sheet.cell(row = row, column = 2).value
                            merge_sheet.cell(row = newRow, column = 3).value = fn
                            merge_sheet.cell(row = newRow, column = 4).value = ln
                            merge_sheet.cell(row = newRow, column = 5).value = email
                            merge_sheet.cell(row = newRow, column = 6).value = merge_sheet.cell(row = row, column = 6).value
                            merge_sheet.cell(row = newRow, column = 7).value = merge_sheet.cell(row = row, column = 7).value
                            merge_sheet.cell(row = newRow, column = 8).value = merge_sheet.cell(row = row, column = 8).value
                            merge_sheet.cell(row = newRow, column = 9).value = merge_sheet.cell(row = row, column = 9).value
                            merge_sheet.cell(row = newRow, column = 10).value = merge_sheet.cell(row = row, column = 10).value
                            merge_sheet.cell(row = newRow, column = 11).value = merge_sheet.cell(row = row, column = 11).value
                    break
                else:
                    continue

            # Save MailMerge
            merge_wb.save("MailMerge.xlsx")
            # Ask user if they'd like to finish and view file, finish and exit, or upload objectives file
            self.empty()


        # Delete empty rows in MailMerge.xlsx
    def empty(self):
        for row in range(2, merge_sheet.max_row + 1):
            print("Deleting empty rows")
            if merge_sheet.cell(row = row, column = 5).value is None:
                for col in range(1, merge_sheet.max_column + 1):
                    merge_sheet.cell(row = row, column = col).value = ""
                merge_wb.save("MailMerge.xlsx")
        merge_wb.save("MailMerge.xlsx")
            
    # def prompt(self):
    #     print("What does user want to do now")
    #     next = QMessageBox()
    #     next.setIcon(QMessageBox.Question)
    #     next.setText("What would you like to do next?")
    #     view = QPushButton.addButton("Finish and open Mail Merge file to view", self)
    #     view.resize(view.sizeHint())
    #     view.move(50,50)
    #     view.clicked.connect(self.close())
    #     close = QPushButton.addButton("Finish and exit", self)
    #     close.resize(close.sizeHint())
    #     close.move(100, 50)
    #     close.clicked.connect(self.close())
    #     contacts = QPushButton.addButton("Upload OSC Master file to update contacts")
    #     contacts.resize(contacts.sizeHint())
    #     contacts.move(50, 200)
    #     contacts.clicked.connect(self.emails)
    #     objectives = QPushButton.addButton("Upload sales objectives workbook to update objectives")
    #     objectives.resize(objectives.sizeHint())
    #     objectives.move(100, 200)
    #     objectives.clicked.connect(self.objectives)
    #     next.exec_()


        # Function to confirm the fields are matched properly to merge    
##        def confirmFields(label_list):
##            print("Confirm fields!")
##            confirm = QMessageBox()
##            confirm.setIcon(QMessageBox.Question)
##            confirm.setText("Are the fields matched to the correct columns?")
##            confirm.setInformativeText("Click 'Show Details' to check the detected merge fields. If the pairings are not correct, you can manually enter the correct columns for each field.")
##            confirm.setDetailedText(label_list[0] + ": OSC Number \n"\
##                                    "-------------------------------- \n"\
##                                    "" + label_list[1] + ": Aero Status \n"\
##                                    "-------------------------------- \n"\
##                                    "" + label_list[2] + ": DBA \n"\
##                                    "-------------------------------- \n"\
##                                    "" + label_list[3] + ": Objective \n"\
##                                    "-------------------------------- \n"\
##                                    "" + label_list[4] + ": MTD \n"\
##                                    "-------------------------------- \n"\
##                                    "" + label_list[5] + ": % of Goal \n"\
##                                    "-------------------------------- \n"\
##                                    "" + label_list[6] + ": PurchasesToGo \n"\
##                                    "-------------------------------- \n"\
##                                    "" + label_list[7] + ": Reward")
##            confirm.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
##            confirm.exec_()
##            if QMessageBox.No:
##                print("Correct fields")
            





    # Start up functions
    # --------------------------------------------------------------------------------------------------------------------------
    def __init__(self):
        super().__init__()
        self.initUI()

        
    def initUI(self):

        # Upload sales objectives button connected to "objectives"
        btn = QPushButton('Upload sales objective file to prep for mail merge', self)
        btn.resize(btn.sizeHint())
        btn.move(50,50)
        btn.clicked.connect(self.objectives)

    # Upload end of month file to calculate rewards
        # Assume OSC qualify = meet monthly goal
        # Assume IRF qualify = Over x amount of sales that month --> 3% of sales; Met goal --> 5% of sales
#        rewardsBtn = QPushButton('Upload end of month file to calculate rewards', self)
#        rewardsBtn.resize(rewardsBtn.sizeHint())
#        rewardsBtn.move(350, 50)

    # Upload OSC Master file button connected to "emails"
        updateBtn = QPushButton('Upload file to update contact information', self)
        updateBtn.resize(updateBtn.sizeHint())
        updateBtn.move(350, 50)
        updateBtn.clicked.connect(self.emails)
        self.setGeometry(100, 100, 800, 500)
        self.setWindowTitle('Orio Email Automation Program')
        self.setWindowIcon(QIcon('logo.png'))        
    
        self.show()


            

if __name__ == '__main__':
    
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())
