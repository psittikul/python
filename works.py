from __future__ import division
import sys
import os
import openpyxl
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import pyqtSignal
from flask import Flask

from tkinter import Tk
from tkinter.filedialog import askopenfilename


THUMBNAIL_SIZE = 128
app = Flask(__name__)


class Example(QWidget):
    progressingSignal = pyqtSignal(float)
    emailProgressingSignal = pyqtSignal(float)
    finishProgressingSignal = pyqtSignal(float)

    def center(self):
        frameGm = self.frameGeometry()
        screen = QApplication.desktop().screenNumber(QApplication.desktop().cursor().pos())
        centerPoint = QApplication.desktop().screenGeometry(screen).center()
        frameGm.moveCenter(centerPoint)
        self.move(frameGm.topLeft())

    def updateProgress(self, value):
        value = value * 100
        self.objectivesProgress.setValue(value)

    def updateEmailsProgress(self, value):
        value = value * 100
        self.emailsProgress.setValue(value)

    def updateFinishProgress(self, value):
        value = value * 100
        self.finishProgress.setValue(value)



    # Function to prompt user for OSC Master file
    def getemailfile(self):
        getOSCMaster = QMessageBox()
        getOSCMaster.setIcon(QMessageBox.Question)
        getOSCMaster.setText("Upload the OSC Master file")
        getOSCMaster.exec_()
        if QMessageBox.Ok:
            self.emails()
        else:
            self.close()

    # Function to actually get the objectives update data
    def objectives(self):

        # Internal helper function to format the currency values with $ and comma separation
        def formatCurrency(amt):
            if len(amt) > 3:
                hundreds = amt[len(amt) - 3:len(amt) + 1]
                thousands = amt[:len(amt) - 3]
                cash = "".join(["$", thousands, ",", hundreds])
                return cash
            else:
                cash = "".join(["$", amt])
                return cash

        # -----------------------------------------------------------------------------------

        Tk().withdraw()
        self.filename = askopenfilename()
        if "Sales" not in str(self.filename):
            confirm = QMessageBox()
            confirm.setIcon(QMessageBox.Warning)
            confirm.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            confirm.setText("This doesn't look like the sales objectives workbook. Are you sure this is it?")
            confirm.exec_()
            if QMessageBox.No:
                Tk().withdraw()
                self.filename = askopenfilename()

        if not self.filename:
            self.close()

        data_wb = openpyxl.load_workbook(self.filename)
        # Go to the correct worksheet
        # Create mail merge file
        merge_wb = openpyxl.Workbook()
        merge_sheet = merge_wb.active
        merge_sheet['A1'] = "OSC Number"
        merge_sheet['B1'] = "DBA"
        merge_sheet['C1'] = "Contact First Name"
        merge_sheet['D1'] = "Contact Last Name"
        merge_sheet['E1'] = "Contact Email"
        merge_sheet['F1'] = "Aero Status"
        merge_sheet['G1'] = "Objective"
        merge_sheet['H1'] = "MTD"
        merge_sheet['I1'] = "% of Goal"
        merge_sheet['J1'] = "Purchases ToGo"
        merge_sheet['K1'] = "Reward"
        merge_wb.save('MailMerge.xlsx')

        # Reactivate the objectives update workbook
        sheet = data_wb.active
        # Cycle through each row. If the row is an OSC, gather this data to later be stored in the array "objectives_update"
        # We'll start populating the mail merge worksheet at row 2
        mergeRow = 2
        # Start at top of file looking for the column headers (taking note of index) to determine the location of each field
        # Then gather those values to be stored in variables for copying over to "Mail Merge"
        for row in range(23, 30):
            for column in range(1, sheet.max_column + 1):
                if sheet.cell(row=row, column=column).value == "Rwd.":
                    rewardColumn = column
                    break
                else:
                    continue
        aeroColumn = rewardColumn - 1
        DBAColumn = rewardColumn + 2
        objectiveColumn = DBAColumn + 2
        PTDColumn = objectiveColumn + 1
        percentColumn = PTDColumn + 1
        PTGColumn = percentColumn + 3



        for row in range(23, sheet.max_row + 1):
            # Send a signal to update the status bar every time we pass another row of the objectives sheet
            signal = row / (sheet.max_row)
            self.progressingSignal.emit(signal)
            if str(sheet.cell(row=row, column=2).value) == "OSC" and type(
                    sheet.cell(row=row, column=rewardColumn).value) == int:
                code = str(sheet.cell(row=row, column=1).value)
                # If the OSC code is negative, take the "-" symbol out of it
                if "-" in code:
                    code = code[1:len(code) + 1]
                else:
                    code = code
                # Reward will be in the column labeled "Rwd."
                rewardCell = sheet.cell(row=row, column=rewardColumn)
                reward = "".join(["$", str(rewardCell.value)])
                # Aero status will be one cell to the left of reward
                aeroCell = sheet.cell(row=row, column=aeroColumn)
                aeroStatus = str(aeroCell.value)
                # DBA will be 2 cells to the right of reward
                DBACell = sheet.cell(row=row, column=DBAColumn)
                DBA = str(DBACell.value)
                # Objective will be 2 cells to the right of DBA
                objectiveCell = sheet.cell(row=row, column=objectiveColumn)
                objective = str(objectiveCell.value)
                objective = formatCurrency(objective)
                # Purchases to date will be 1 cell to the right of objective
                PTDCell = sheet.cell(row=row, column=PTDColumn)
                PTD = str(PTDCell.value)
                PTD = formatCurrency(PTD)
                # % of objective will be 1 cell to the right of PTD
                percentCell = sheet.cell(row=row, column=percentColumn)
                percent = int(percentCell.value * 100)
                percent = "".join([str(percent), "%"])
                # Purchases to go will be 3 cells to the right of % of objective
                PTGCell = sheet.cell(row=row, column=PTGColumn)
                PTG = str(PTGCell.value)
                PTG = formatCurrency(PTG)
                # Now that you've gotten the information, copy it over to Mail Merge
                merge_sheet = merge_wb.active
                # Go through all the info columns in Mail Merge and copy over the values
                merge_sheet.cell(row=mergeRow, column=1).value = code
                merge_sheet.cell(row=mergeRow, column=2).value = DBA
                merge_sheet.cell(row=mergeRow, column=6).value = aeroStatus
                merge_sheet.cell(row=mergeRow, column=7).value = objective
                merge_sheet.cell(row=mergeRow, column=8).value = PTD
                merge_sheet.cell(row=mergeRow, column=9).value = percent
                merge_sheet.cell(row=mergeRow, column=10).value = PTG
                merge_sheet.cell(row=mergeRow, column=11).value = reward
                # Now that we've copied the information to this row, get the next merge row ready and reactivate the objectives workbook and worksheet
                merge_wb.save("MailMerge.xlsx")
                mergeRow = mergeRow + 1
                sheet = data_wb.active
            else:
                continue
            merge_sheet = merge_wb.active
        merge_wb.save("MailMerge.xlsx")
        # Now prompt the user to upload the OSC Master file so we can add the emails
        self.getemailfile()

    # Function to ask for OSC Master upload and copy those emails to the Mail Merge file
    def emails(self):
        Tk().withdraw()
        self.filename = askopenfilename()
        if "OSC" not in str(self.filename):
            confirm = QMessageBox()
            confirm.setIcon(QMessageBox.Warning)
            confirm.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            confirm.setText("This doesn't look like the OSC Master workbook. Are you sure this is it?")
            confirm.exec_()
            if QMessageBox.No:
                Tk().withdraw()
                self.filename = askopenfilename()

        if not self.filename:
            self.close()
        OSC_wb = openpyxl.load_workbook(self.filename)
        # Go to the correct worksheet
        OSC_sheet = OSC_wb.active
        # Cycle through each row. If the row is an OSC, gather this data to later be stored in the array "objectives_update"
        # We'll start populating the mail merge worksheet at row 2
        # Start at header row of OSC Master to determine the column each manager's email is in
        for column in range(1, OSC_sheet.max_column + 1):
            if str(OSC_sheet.cell(row=1, column=column).value) == "OSC":
                code_column = column
            if str(OSC_sheet.cell(row=1, column=column).value) == "General Manager Name":
                GM_name_column = column
            if str(OSC_sheet.cell(row=1, column=column).value) == "Email (1)":
                GM_email_column = column
            if str(OSC_sheet.cell(row=1, column=column).value) == "Parts Manager Name":
                PM_name_column = column
            if str(OSC_sheet.cell(row=1, column=column).value) == "Email (2)":
                PM_email_column = column
            if str(OSC_sheet.cell(row=1, column=column).value) == "Service Manager Name":
                SM_name_column = column
            if str(OSC_sheet.cell(row=1, column=column).value) == "Email (3)":
                SM_email_column = column
                break
        # Now that you've determined the location of each field, go through every row in the sheet to get contact info
        # But skip any emails that have already been copied or are blank
        for row in range(2, OSC_sheet.max_row + 1):
            emailSignal = row / OSC_sheet.max_row
            self.emailProgressingSignal.emit(emailSignal)
            if "Active" in str(OSC_sheet.cell(row=row, column=1).value):
                OSC_names = []
                OSC_emails = []
                PM_email = str(OSC_sheet.cell(row=row, column=PM_email_column).value)
                GM_email = str(OSC_sheet.cell(row=row, column=GM_email_column).value)
                SM_email = str(OSC_sheet.cell(row=row, column=SM_email_column).value)
                code = str(OSC_sheet.cell(row=row, column=code_column).value)
                if "@" in GM_email:
                    GM_name = str(OSC_sheet.cell(row=row, column=GM_name_column).value)
                    OSC_names.append(GM_name)
                    OSC_emails.append(GM_email)
                if "@" in PM_email and PM_email != GM_email and PM_email != SM_email:
                    PM_name = str(OSC_sheet.cell(row=row, column=PM_name_column).value)
                    OSC_names.append(PM_name)
                    OSC_emails.append(PM_email)
                if "@" in SM_email and SM_email != PM_email and SM_email != GM_email:
                    SM_name = str(OSC_sheet.cell(row=row, column=SM_name_column).value)
                    OSC_names.append(SM_name)
                    OSC_emails.append(SM_email)
            else:
                continue
            # Now that you've gotten the information, copy it over to Mail Merge
            mergeFilePath = str(os.getcwd()) + "\MailMerge.xlsx"
            merge_wb = openpyxl.load_workbook(mergeFilePath)
            merge_sheet = merge_wb.active
            for row in range(2, merge_sheet.max_row + 1):
                value_code = str(merge_sheet.cell(row=row, column=1).value) == code
                if value_code:
                    name = OSC_names[0]
                    namePieces = name.split()
                    if len(namePieces) == 2:
                        fn = namePieces[0]
                        ln = namePieces[1]
                    elif len(namePieces) == 3:
                        fn = "".join([namePieces[0], " ", namePieces[1]])
                        ln = namePieces[2]
                    elif len(namePieces) == 4:
                        fn = "".join([namePieces[0], " ", namePieces[1]])
                        ln = "".join([namePieces[2], " ", namePieces[3]])
                    email = OSC_emails[0]
                    merge_sheet.cell(row=row, column=3).value = fn
                    merge_sheet.cell(row=row, column=4).value = ln
                    merge_sheet.cell(row=row, column=5).value = email
                    if len(OSC_emails) > 1:
                        for em in OSC_emails:
                            curr_index = OSC_emails.index(em) + 1
                            if curr_index > len(OSC_emails) - 1:
                                break
                            newRow = merge_sheet.max_row + 1
                            name = OSC_names[curr_index]
                            namePieces = name.split()
                            if len(namePieces) == 2:
                                fn = namePieces[0]
                                ln = namePieces[1]
                            elif len(namePieces) == 3:
                                fn = "".join([namePieces[0], " ", namePieces[1]])
                                ln = namePieces[2]
                            elif len(namePieces) == 4:
                                fn = "".join([namePieces[0], " ", namePieces[1]])
                                ln = "".join([namePieces[2], " ", namePieces[3]])
                            email = OSC_emails[curr_index]
                            merge_sheet.cell(row=newRow, column=1).value = code
                            merge_sheet.cell(row=newRow, column=2).value = merge_sheet.cell(row=row, column=2).value
                            merge_sheet.cell(row=newRow, column=3).value = fn
                            merge_sheet.cell(row=newRow, column=4).value = ln
                            merge_sheet.cell(row=newRow, column=5).value = email
                            merge_sheet.cell(row=newRow, column=6).value = merge_sheet.cell(row=row, column=6).value
                            merge_sheet.cell(row=newRow, column=7).value = merge_sheet.cell(row=row, column=7).value
                            merge_sheet.cell(row=newRow, column=8).value = merge_sheet.cell(row=row, column=8).value
                            merge_sheet.cell(row=newRow, column=9).value = merge_sheet.cell(row=row, column=9).value
                            merge_sheet.cell(row=newRow, column=10).value = merge_sheet.cell(row=row, column=10).value
                            merge_sheet.cell(row=newRow, column=11).value = merge_sheet.cell(row=row, column=11).value
                    continue
                else:
                    continue
            # Save MailMerge
            merge_wb.save("MailMerge.xlsx")
        merge_wb.save("MailMerge.xlsx")
        # Ask user if they'd like to finish and view file or finish and exit
        self.empty()


        # Delete empty rows in MailMerge.xlsx

    def empty(self):
        mergeFilePath = "".join([str(os.getcwd()), "\MailMerge.xlsx"])
        merge_wb = openpyxl.load_workbook(mergeFilePath)
        merge_sheet = merge_wb.active
        for row in range(2, merge_sheet.max_row + 1):
            signal = row / merge_sheet.max_row
            self.finishProgressingSignal.emit(signal)
            if merge_sheet.cell(row=row, column=5).value is None:
                for col in range(1, merge_sheet.max_column + 1):
                    merge_sheet.cell(row=row, column=col).value = ""
                merge_wb.save("MailMerge.xlsx")
        merge_wb.save("MailMerge.xlsx")
        self.finalprompt()

    # Start up functions
    # --------------------------------------------------------------------------------------------------------------------------
    def finalprompt(self):
        finalprompt = QMessageBox()
        finalprompt.setIcon(QMessageBox.Question)
        finalprompt.setText("Open and view completed Mail Merge file?")
        finalprompt.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        finalprompt.exec_()
        if QMessageBox.Yes:
            os.startfile("".join([str(os.getcwd()), "\MailMerge.xlsx"]))
            self.close()
        else:
            self.close()

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        # Upload sales objectives button connected to "objectives"
        btn = QPushButton('Upload objectives file to begin mail merge process for weekly OSC objectives update', self)
        btn.resize(btn.sizeHint())
        btn.move(100, 40)
        btn.clicked.connect(self.objectives)

        self.objectivesLabel = QLabel(self)
        self.emailsLabel = QLabel(self)
        self.finishLabel = QLabel(self)
        self.objectivesLabel.setText("Uploading objectives...")
        self.emailsLabel.setText("Uploading OSC emails...")
        self.finishLabel.setText("Finishing and cleaning up MailMerge.xlsx...")
        self.objectivesLabel.setGeometry(200, 75, 250, 20)
        self.emailsLabel.setGeometry(200, 125, 250, 20)
        self.finishLabel.setGeometry(200, 175, 250, 20)

        # Upload end of month file to calculate rewards
        # Assume OSC qualify = meet monthly goal
        # Assume IRF qualify = Over x amount of sales that month --> 3% of sales; Met goal --> 5% of sales
        #        rewardsBtn = QPushButton('Upload end of month file to calculate rewards', self)
        #        rewardsBtn.resize(rewardsBtn.sizeHint())
        #        rewardsBtn.move(350, 50)
        self.objectivesProgress = QProgressBar(self)
        self.objectivesProgress.setGeometry(200, 100, 250, 20)
        self.emailsProgress = QProgressBar(self)
        self.emailsProgress.setGeometry(200, 150, 250, 20)
        self.finishProgress = QProgressBar(self)
        self.finishProgress.setGeometry(200, 200, 250, 20)
        self.progressingSignal.connect(self.updateProgress)
        self.emailProgressingSignal.connect(self.updateEmailsProgress)
        self.finishProgressingSignal.connect(self.updateFinishProgress)

        self.setGeometry(100, 100, 650, 500)
        self.setWindowTitle('Orio Email Mail Merge Automation Program')
        l = QPixmap("email-icon.png")
        if l.height() > l.width():
            l = l.scaledToWidth(THUMBNAIL_SIZE)
        else:
            l = l.scaledToHeight(THUMBNAIL_SIZE)
        l = l.copy(0, 0, THUMBNAIL_SIZE, THUMBNAIL_SIZE)
        self.setWindowIcon(QIcon(l))
        self.center()
        self.setGeometry(self.frameGeometry())
        self.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())