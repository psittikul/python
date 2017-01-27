# reads Excel file of client log, arranges data correctly for import into Access database

import openpyxl, pprint, os
path = "C:/Users/Patricia Sittikul/SharePoint/Info LincLV/Shared/1 Clients/Active/CT and DC/Stillson, Eli"
os.chdir(path)
print(os.getcwd())

wb = openpyxl.load_workbook('Honerkamp Smith Client Log.xlsx')
print ('Opening workbook...')
sheet = wb.get_sheet_by_name('Sheet1')
w, h = sheet.max_row + 1, 11
Matrix = [[0 for x in range(w)] for y in range(h)]


""" Make sure all client log fields are correctly formatted,
    and if they are, move them to the file to import into Access.
"""



for rowNum in range(2, int(len(Matrix)) + 1):
    initials = sheet.cell(row = rowNum, column=1).value
    Matrix[0][rowNum] = initials
    date = sheet.cell(row = rowNum, column = 2).value
    Matrix[1][rowNum] = date
    face = sheet.cell(row = rowNum, column = 3).value
    Matrix[2][rowNum] = face
    phone = sheet.cell(row = rowNum, column = 4).value
    Matrix[3][rowNum] = phone
    email = sheet.cell(row = rowNum, column = 5).value
    Matrix[4][rowNum] = email
    category = sheet.cell(row=rowNum, column=6).value
    Matrix[5][rowNum] = category
    print(category)
    """error1 = str('Interviwiews')
    error2 = str('Resources_Job Search')
    error3 = str('Resources_Coaching')
    error4 = str('Individual Connections')
    error5 = str('Resources_Events')
    if error1 in str(category):
        category = 'Informational Interviews'
        sheet.cell(row=rowNum, column=6).value = category
    elif error2 in str(category):
        category = 'Resources - Job Search'
        sheet.cell(row=rowNum, column=6).value = category
    elif error3 in str(category):
        category = 'Resources - Coaching'
        sheet.cell(row=rowNum, column=6).value = category
    elif error4 in str(category):
        category = 'Individual Connections'
    elif error5 in str(category):
        category = 'Resources - Events'
        sheet.cell(row=rowNum, column=6).value = category
    else: category = category
    wb.save('Honerkamp Smith Client Log.xlsx') """
    outcome = sheet.cell(row = rowNum, column=7).value
    Matrix[6][rowNum] = outcome
    resources = sheet.cell(row = rowNum, column = 8).value
    Matrix[7][rowNum] = resources
    numResources = sheet.cell(row = rowNum, column = 9).value
    Matrix[8][rowNum] = numResources
    time = sheet.cell(row = rowNum, column = 10).value
    Matrix[9][rowNum] = time
    clientID = 49
    Matrix[10][rowNum] = 49
print(len(Matrix))
print(os.getcwd())
resultWB = openpyxl.load_workbook('honerkampclientimport.xlsx')
sheet = resultWB.get_sheet_by_name('Sheet1')
print('Writing results...')
sheet.cell(1,1).value = 'ID'
sheet.cell(1, 2).value = 'Initials'
sheet.cell(1, 3).value = 'Title'
sheet.cell(1, 4).value = 'LogID'
sheet.cell(1, 5).value = 'Date'
sheet.cell(1, 6).value = 'Face-to-Face'
sheet.cell(1, 7).value = 'Phone'
sheet.cell(1, 8).value = 'Email'
sheet.cell(1, 9).value = 'Service'
sheet.cell(1, 10).value = 'Category'
sheet.cell(1, 11).value = 'Outcome'
sheet.cell(1, 12).value = 'Resources Offered'
sheet.cell(1, 13).value = '# of Resources Offered'
sheet.cell(1, 14).value = 'TimeSpent'
sheet.cell(1, 15).value = 'ClientID'
"""for num in range(2, 12):
    sheet[num, 2].value = Matrix[0][num]
    sheet[num, 5].value=Matrix[1][num]"""
"""sheet.cell(index, column=5).value = Matrix[1][num]
    sheet.cell(index, column=6).value = Matrix[2][rowNum]
    sheet.cell(row=index, column=7).value = Matrix[3][rowNum]
    sheet.cell(row=index, column=8).value = Matrix[4][rowNum]
    sheet.cell(row=index, column=10).value = Matrix[5][rowNum]
    sheet.cell(row=index, column=11).value = Matrix[6][rowNum]
    sheet.cell(row=index, column=12).value = Matrix[7][rowNum]
    sheet.cell(row=index, column=13).value = Matrix[8][rowNum]
    sheet.cell(row=index, column=14).value = Matrix[9][rowNum]
    sheet.cell(row=index, column=15).value = Matrix[10][rowNum]"""
resultWB.save('honerkampclientimport.xlsx')
        
